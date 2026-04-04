"""Tests for API clients (mocked HTTP), Excel/CSV output, ValidatorStatsCache, and gather_events."""
import os
import sys
import json
import pytest
from decimal import Decimal
from datetime import datetime, timezone
from unittest.mock import patch, MagicMock, PropertyMock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from staking_tax_report import (
    EtherscanClient,
    BeaconClient,
    BeaconchainClient,
    CryptoCompareClient,
    ValidatorStatsCache,
    gather_events,
    write_excel_bytes,
    write_csv_bytes,
    _prepare_events,
    GWEI_PER_ETH,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _mock_response(json_data, status_code=200):
    resp = MagicMock()
    resp.status_code = status_code
    resp.json.return_value = json_data
    resp.raise_for_status.return_value = None
    if status_code >= 400:
        from requests.exceptions import HTTPError
        resp.raise_for_status.side_effect = HTTPError(response=resp)
    return resp


# ---------------------------------------------------------------------------
# EtherscanClient
# ---------------------------------------------------------------------------

class TestEtherscanClient:
    def test_get_beacon_withdrawals_basic(self):
        """get_beacon_withdrawals returns parsed withdrawal events."""
        client = EtherscanClient("TESTKEY", log_fn=lambda x: None)
        fake_result = {
            "status": "1",
            "message": "OK",
            "result": [
                {
                    "withdrawalIndex": "1",
                    "validatorIndex": "100",
                    "amount": "1000000000",  # 1 ETH in Gwei
                    "timestamp": "1700000000",
                    "blockNumber": "18000000",
                },
            ],
        }
        with patch.object(client, "_get", return_value=fake_result):
            results = client.get_beacon_withdrawals("0x" + "a" * 40, 1699000000, 1701000000)
        assert len(results) == 1
        assert results[0]["validator"] == 100
        assert results[0]["amount_eth"] == Decimal("1000000000") / GWEI_PER_ETH
        assert results[0]["type"] == "withdrawal"

    def test_get_beacon_withdrawals_empty(self):
        """No withdrawals returns empty list."""
        client = EtherscanClient("TESTKEY", log_fn=lambda x: None)
        fake_result = {"status": "1", "message": "OK", "result": []}
        with patch.object(client, "_get", return_value=fake_result):
            results = client.get_beacon_withdrawals("0x" + "a" * 40, 1699000000, 1701000000)
        assert results == []

    def test_get_beacon_withdrawals_no_transactions(self):
        """'No transactions found' message returns empty list."""
        client = EtherscanClient("TESTKEY", log_fn=lambda x: None)
        fake_result = {"status": "0", "message": "No transactions found", "result": []}
        with patch.object(client, "_get", return_value=fake_result):
            results = client.get_beacon_withdrawals("0x" + "a" * 40, 1699000000, 1701000000)
        assert results == []

    def test_get_beacon_withdrawals_filters_by_timestamp(self):
        """Withdrawals outside the date range are filtered out."""
        client = EtherscanClient("TESTKEY", log_fn=lambda x: None)
        fake_result = {
            "status": "1",
            "message": "OK",
            "result": [
                {"withdrawalIndex": "1", "validatorIndex": "100", "amount": "500000000",
                 "timestamp": "1699500000", "blockNumber": "18000000"},
                {"withdrawalIndex": "2", "validatorIndex": "100", "amount": "500000000",
                 "timestamp": "1702000000", "blockNumber": "18000001"},  # outside range
            ],
        }
        with patch.object(client, "_get", return_value=fake_result):
            results = client.get_beacon_withdrawals("0x" + "a" * 40, 1699000000, 1701000000)
        assert len(results) == 1
        assert results[0]["timestamp"] == 1699500000

    def test_fallback_not_used_for_beacon_withdrawals(self):
        """txsBeaconWithdrawal is Etherscan-only, should not fall back."""
        client = EtherscanClient("TESTKEY", log_fn=lambda x: None)
        rate_limit_resp = {"status": "0", "message": "NOTOK", "result": "Max rate limit reached"}
        with patch.object(client, "_get", return_value=rate_limit_resp):
            # Should get empty result (status != "1" and no "No transactions found")
            results = client.get_beacon_withdrawals("0x" + "a" * 40, 1699000000, 1701000000)
        assert results == []

    def test_api_key_masked_in_errors(self):
        """API keys should not appear in error messages."""
        client = EtherscanClient("MYSECRETKEY123", log_fn=lambda x: None)
        from requests.exceptions import ConnectionError
        with patch.object(client, "_throttle"), \
             patch.object(client.session, "get", side_effect=ConnectionError("https://api.etherscan.io?apikey=MYSECRETKEY123")):
            with pytest.raises(RuntimeError) as exc_info:
                client.get_beacon_withdrawals("0x" + "a" * 40, 1699000000, 1701000000)
            assert "MYSECRETKEY123" not in str(exc_info.value)


# ---------------------------------------------------------------------------
# BeaconClient
# ---------------------------------------------------------------------------

class TestBeaconClient:
    def test_get_validator_info_basic(self):
        """Returns parsed validator info for 0x01 validator."""
        client = BeaconClient(log_fn=lambda x: None)
        fake_data = {
            "data": {
                "validator": {
                    "pubkey": "0x" + "ab" * 48,
                    "withdrawal_credentials": "0x01" + "00" * 11 + "aa" * 20,
                    "effective_balance": "32000000000",
                },
                "index": "12345",
                "status": "active_ongoing",
                "balance": "32500000000",
            }
        }
        with patch.object(client.session, "get", return_value=_mock_response(fake_data)):
            info = client.get_validator_info("12345")
        assert info["validator_index"] == 12345
        assert info["credential_type"] == "0x01"
        assert info["status"] == "active_ongoing"
        assert info["balance_eth"] == Decimal("32500000000") / GWEI_PER_ETH
        assert info["withdrawal_address"].startswith("0x")

    def test_get_validator_info_0x02(self):
        """0x02 compounding validator detected correctly."""
        client = BeaconClient(log_fn=lambda x: None)
        fake_data = {
            "data": {
                "validator": {
                    "pubkey": "0x" + "cd" * 48,
                    "withdrawal_credentials": "0x02" + "00" * 11 + "bb" * 20,
                    "effective_balance": "2048000000000",
                },
                "index": "99999",
                "status": "active_ongoing",
                "balance": "2100000000000",
            }
        }
        with patch.object(client.session, "get", return_value=_mock_response(fake_data)):
            info = client.get_validator_info("99999")
        assert info["credential_type"] == "0x02"

    def test_get_validator_info_bls_raises(self):
        """BLS (0x00) credentials should raise RuntimeError."""
        client = BeaconClient(log_fn=lambda x: None)
        fake_data = {
            "data": {
                "validator": {
                    "pubkey": "0x" + "ee" * 48,
                    "withdrawal_credentials": "0x00" + "ff" * 31,
                    "effective_balance": "32000000000",
                },
                "index": "11111",
                "status": "active_ongoing",
                "balance": "32000000000",
            }
        }
        with patch.object(client.session, "get", return_value=_mock_response(fake_data)):
            with pytest.raises(RuntimeError, match="BLS"):
                client.get_validator_info("11111")

    def test_get_validator_info_not_found(self):
        """Non-existent validator raises RuntimeError."""
        client = BeaconClient(log_fn=lambda x: None)
        resp = MagicMock()
        resp.status_code = 404
        with patch.object(client.session, "get", return_value=resp):
            with pytest.raises(RuntimeError, match="not found"):
                client.get_validator_info("99999999")


# ---------------------------------------------------------------------------
# ValidatorStatsCache
# ---------------------------------------------------------------------------

class TestValidatorStatsCache:
    def test_store_and_retrieve(self, tmp_path):
        db = str(tmp_path / "cache.db")
        cache = ValidatorStatsCache(db)
        stats = [
            {"date": "2024-01-01", "start_balance": 32000000000, "end_balance": 32001000000,
             "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 1000000},
            {"date": "2024-01-02", "start_balance": 32001000000, "end_balance": 32002500000,
             "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 1500000},
        ]
        cache.store(100, stats)
        result = cache.get_cached(100, "2024-01-01", "2024-01-02")
        assert len(result) == 2
        assert result[0]["date"] == "2024-01-01"
        assert result[0]["reward_gwei"] == 1000000
        cache.close()

    def test_get_cached_empty(self, tmp_path):
        db = str(tmp_path / "cache.db")
        cache = ValidatorStatsCache(db)
        result = cache.get_cached(999, "2024-01-01", "2024-01-31")
        assert result == []
        cache.close()

    def test_get_cached_dates(self, tmp_path):
        db = str(tmp_path / "cache.db")
        cache = ValidatorStatsCache(db)
        stats = [
            {"date": "2024-01-01", "start_balance": 0, "end_balance": 0,
             "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 100},
            {"date": "2024-01-03", "start_balance": 0, "end_balance": 0,
             "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 200},
        ]
        cache.store(50, stats)
        dates = cache.get_cached_dates(50, "2024-01-01", "2024-01-05")
        assert dates == {"2024-01-01", "2024-01-03"}
        cache.close()

    def test_different_validators_isolated(self, tmp_path):
        db = str(tmp_path / "cache.db")
        cache = ValidatorStatsCache(db)
        stats1 = [{"date": "2024-01-01", "start_balance": 0, "end_balance": 0,
                    "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 100}]
        stats2 = [{"date": "2024-01-01", "start_balance": 0, "end_balance": 0,
                    "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 999}]
        cache.store(1, stats1)
        cache.store(2, stats2)
        r1 = cache.get_cached(1, "2024-01-01", "2024-01-01")
        r2 = cache.get_cached(2, "2024-01-01", "2024-01-01")
        assert r1[0]["reward_gwei"] == 100
        assert r2[0]["reward_gwei"] == 999
        cache.close()

    def test_duplicate_insert_ignored(self, tmp_path):
        db = str(tmp_path / "cache.db")
        cache = ValidatorStatsCache(db)
        stat = {"date": "2024-01-01", "start_balance": 0, "end_balance": 0,
                "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 100}
        cache.store(1, [stat])
        cache.store(1, [stat])  # duplicate
        result = cache.get_cached(1, "2024-01-01", "2024-01-01")
        assert len(result) == 1
        cache.close()


# ---------------------------------------------------------------------------
# BeaconchainClient
# ---------------------------------------------------------------------------

class TestBeaconchainClient:
    def test_fetch_stats_parses_null_values(self):
        """Null values in beaconcha.in response should be treated as 0."""
        client = BeaconchainClient(api_key="test", cache_path=":memory:", log_fn=lambda x: None)
        fake_data = {
            "status": "OK",
            "data": [
                {
                    "day_start": "2024-01-01T00:00:00Z",
                    "start_balance": 32000000000,
                    "end_balance": 32001000000,
                    "deposits_amount": None,  # null
                    "withdrawals_amount": None,  # null
                },
            ],
        }
        with patch.object(client, "_get", return_value=fake_data):
            results = client._fetch_stats(100)
        assert len(results) == 1
        assert results[0]["deposits_amount"] == 0
        assert results[0]["withdrawals_amount"] == 0
        assert results[0]["reward_gwei"] == 1000000  # end - start
        client.close()

    def test_fetch_stats_empty_response(self):
        """Empty data from beaconcha.in returns empty list."""
        client = BeaconchainClient(api_key="test", cache_path=":memory:", log_fn=lambda x: None)
        with patch.object(client, "_get", return_value={"status": "OK", "data": []}):
            results = client._fetch_stats(100)
        assert results == []
        client.close()

    def test_fetch_stats_not_ok_status(self):
        """Non-OK status returns empty list."""
        client = BeaconchainClient(api_key="test", cache_path=":memory:", log_fn=lambda x: None)
        with patch.object(client, "_get", return_value={"status": "ERROR", "message": "fail"}):
            results = client._fetch_stats(100)
        assert results == []
        client.close()

    def test_get_daily_rewards_batch_uses_cache(self, tmp_path):
        """Cached data should not trigger API calls."""
        db = str(tmp_path / "cache.db")
        client = BeaconchainClient(api_key="test", cache_path=db, log_fn=lambda x: None)
        # Pre-populate cache
        client.cache.store(100, [
            {"date": "2024-01-01", "start_balance": 32000000000, "end_balance": 32001000000,
             "deposits_amount": 0, "withdrawals_amount": 0, "reward_gwei": 1000000},
        ])
        with patch.object(client, "_fetch_stats") as mock_fetch:
            result = client.get_daily_rewards_batch([100], "2024-01-01", "2024-01-01")
            mock_fetch.assert_not_called()
        assert 100 in result
        assert len(result[100]) == 1
        assert result[100][0]["reward_gwei"] == 1000000
        client.close()

    def test_reward_formula_correct(self):
        """reward = (end - start) - deposits + withdrawals."""
        client = BeaconchainClient(api_key="test", cache_path=":memory:", log_fn=lambda x: None)
        fake_data = {
            "status": "OK",
            "data": [
                {
                    "day_start": "2024-01-01T00:00:00Z",
                    "start_balance": 32000000000,
                    "end_balance": 32100000000,
                    "deposits_amount": 50000000,  # 0.05 ETH deposit
                    "withdrawals_amount": 20000000,  # 0.02 ETH withdrawn
                },
            ],
        }
        with patch.object(client, "_get", return_value=fake_data):
            results = client._fetch_stats(100)
        # reward = (32100000000 - 32000000000) - 50000000 + 20000000 = 70000000
        assert results[0]["reward_gwei"] == 70000000
        client.close()


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

class TestCSVOutput:
    def _make_events(self):
        return [
            {"timestamp": 1700000000, "amount_eth": Decimal("1.5"), "type": "withdrawal", "validator": 100},
            {"timestamp": 1700100000, "amount_eth": Decimal("0.01"), "type": "block_reward", "validator": 100},
        ]

    def test_csv_bytes_returns_bytes(self):
        events = self._make_events()
        prices = [(1700000000, Decimal("2000")), (1700100000, Decimal("2010"))]
        result = write_csv_bytes(events, prices, currency="EUR")
        assert isinstance(result, bytes)

    def test_csv_contains_header(self):
        events = self._make_events()
        prices = [(1700000000, Decimal("2000"))]
        result = write_csv_bytes(events, prices, currency="EUR").decode("utf-8")
        assert "Date" in result
        assert "Amount [ETH]" in result
        assert "Price [EUR/ETH]" in result
        assert "Value [EUR]" in result

    def test_csv_contains_events(self):
        events = self._make_events()
        prices = [(1700000000, Decimal("2000"))]
        result = write_csv_bytes(events, prices, currency="USD").decode("utf-8")
        assert "withdrawal" in result
        assert "block_reward" in result

    def test_csv_row_count(self):
        events = self._make_events()
        prices = [(1700000000, Decimal("2000"))]
        result = write_csv_bytes(events, prices).decode("utf-8")
        lines = [l for l in result.strip().split("\n") if l]
        assert len(lines) == 3  # header + 2 events

    def test_csv_empty_events(self):
        result = write_csv_bytes([], [(1000, Decimal("100"))]).decode("utf-8")
        lines = [l for l in result.strip().split("\n") if l]
        assert len(lines) == 1  # header only


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

class TestExcelOutput:
    def _make_events(self):
        return [
            {"timestamp": 1700000000, "amount_eth": Decimal("1.5"), "type": "withdrawal", "validator": 100},
            {"timestamp": 1700100000, "amount_eth": Decimal("0.01"), "type": "block_reward", "validator": 100},
        ]

    def test_excel_bytes_returns_bytes(self):
        events = self._make_events()
        prices = [(1700000000, Decimal("2000")), (1700100000, Decimal("2010"))]
        result = write_excel_bytes(events, prices, currency="EUR")
        assert isinstance(result, bytes)
        # Verify it's a valid xlsx (starts with PK zip signature)
        assert result[:2] == b"PK"

    def test_excel_with_validator_info(self):
        events = self._make_events()
        prices = [(1700000000, Decimal("2000"))]
        vi = [{
            "index": 100, "name": "test", "credential_type": "0x01",
            "status": "active_ongoing", "balance_eth": Decimal("32"),
            "effective_balance_eth": Decimal("32"),
            "withdrawn_eth": Decimal("1.5"), "withdrawn_income_eth": Decimal("1.5"),
        }]
        result = write_excel_bytes(events, prices, currency="EUR", validator_info=vi)
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_excel_with_accrual_events(self):
        """Excel should handle accrual events without crashing."""
        events = [
            {"timestamp": 1700000000, "amount_eth": Decimal("0.003"), "type": "accrual", "validator": "all"},
            {"timestamp": 1700100000, "amount_eth": Decimal("0.01"), "type": "block_reward", "validator": 100},
        ]
        prices = [(1700000000, Decimal("2000")), (1700100000, Decimal("2010"))]
        result = write_excel_bytes(events, prices, currency="USD")
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_excel_empty_events(self):
        """Empty events should still produce valid Excel."""
        result = write_excel_bytes([], [(1000, Decimal("100"))], currency="EUR")
        assert isinstance(result, bytes)
        assert result[:2] == b"PK"

    def test_excel_summary_number_format_fix(self):
        """Verify the number format fix: ETH columns get 9 decimals, fiat get 2."""
        import openpyxl
        import io
        events = [
            {"timestamp": 1700000000, "amount_eth": Decimal("0.003"), "type": "accrual", "validator": "all"},
            {"timestamp": 1700100000, "amount_eth": Decimal("0.5"), "type": "withdrawal", "validator": 100},
            {"timestamp": 1700200000, "amount_eth": Decimal("0.01"), "type": "block_reward", "validator": 100},
        ]
        prices = [(1700000000, Decimal("2000")), (1700100000, Decimal("2010")), (1700200000, Decimal("2020"))]
        result = write_excel_bytes(events, prices, currency="EUR")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        # Row 2 is first data row; columns after "Month" are pairs of ETH/fiat
        # Check that ETH columns have 9 decimals and fiat have 2
        header = [c.value for c in ws[1]]
        for col_idx, h in enumerate(header, 1):
            if col_idx == 1:
                continue  # "Month" column
            cell = ws.cell(row=2, column=col_idx)
            if cell.value is not None:
                if "[ETH]" in str(h):
                    assert cell.number_format == "0.000000000", f"Column {h} should be ETH format"
                elif "[EUR]" in str(h):
                    assert cell.number_format == "#,##0.00", f"Column {h} should be fiat format"


# ---------------------------------------------------------------------------
# gather_events (mocked dependencies)
# ---------------------------------------------------------------------------

class TestGatherEvents:
    def _make_account(self, **overrides):
        base = {
            "name": "test",
            "withdrawal_address": "0x" + "aa" * 20,
        }
        base.update(overrides)
        return base

    def test_basic_withdrawal_mode(self):
        """gather_events returns withdrawal events in withdrawal mode."""
        account = self._make_account()
        mock_etherscan = MagicMock(spec=EtherscanClient)
        mock_etherscan.get_beacon_withdrawals.return_value = [
            {"validator": 100, "amount_eth": Decimal("0.5"), "type": "withdrawal",
             "timestamp": 1700000000, "blockNumber": 18000000},
        ]
        mock_etherscan.get_execution_rewards.return_value = []

        events = gather_events(
            account, mock_etherscan, 1699000000, 1701000000,
            log_fn=lambda x: None, reporting_mode="withdrawal",
            cache_path=":memory:",
        )
        assert len(events) == 1
        assert events[0]["type"] == "withdrawal"

    def test_no_withdrawal_address_raises(self):
        """Missing withdrawal address should raise RuntimeError."""
        account = {"name": "test"}
        mock_etherscan = MagicMock(spec=EtherscanClient)
        with pytest.raises(RuntimeError, match="withdrawal"):
            gather_events(
                account, mock_etherscan, 1699000000, 1701000000,
                log_fn=lambda x: None, cache_path=":memory:",
            )

    def test_empty_withdrawal_address_raises(self):
        """Empty string withdrawal address should raise RuntimeError."""
        account = self._make_account(withdrawal_address="")
        mock_etherscan = MagicMock(spec=EtherscanClient)
        with pytest.raises(RuntimeError, match="withdrawal"):
            gather_events(
                account, mock_etherscan, 1699000000, 1701000000,
                log_fn=lambda x: None, cache_path=":memory:",
            )

    def test_progress_callback_called(self):
        """Progress function should be called during execution."""
        account = self._make_account()
        mock_etherscan = MagicMock(spec=EtherscanClient)
        mock_etherscan.get_beacon_withdrawals.return_value = []
        mock_etherscan.get_execution_rewards.return_value = []

        progress_values = []
        gather_events(
            account, mock_etherscan, 1699000000, 1701000000,
            log_fn=lambda x: None, reporting_mode="withdrawal",
            cache_path=":memory:",
            progress_fn=lambda pct: progress_values.append(pct),
            progress_range=(0, 100),
        )
        assert len(progress_values) > 0

    def test_validator_filter(self):
        """Events should be filtered by validator_indices when specified."""
        account = self._make_account(validator_indices=[100])
        mock_etherscan = MagicMock(spec=EtherscanClient)
        mock_etherscan.get_beacon_withdrawals.return_value = [
            {"validator": 100, "amount_eth": Decimal("0.5"), "type": "withdrawal",
             "timestamp": 1700000000, "blockNumber": 18000000},
            {"validator": 200, "amount_eth": Decimal("0.3"), "type": "withdrawal",
             "timestamp": 1700000001, "blockNumber": 18000001},
        ]
        mock_etherscan.get_execution_rewards.return_value = []

        events = gather_events(
            account, mock_etherscan, 1699000000, 1701000000,
            log_fn=lambda x: None, reporting_mode="withdrawal",
            cache_path=":memory:",
        )
        assert len(events) == 1
        assert events[0]["validator"] == 100


# ---------------------------------------------------------------------------
# Download endpoint (app.py fix verification)
# ---------------------------------------------------------------------------

class TestDownloadEndpoint:
    @pytest.fixture
    def client(self):
        from app import app, limiter
        app.config["TESTING"] = True
        limiter.enabled = False
        with app.test_client() as c:
            yield c
        limiter.enabled = True

    def test_download_nonexistent_job(self, client):
        resp = client.get("/download/nonexistent-id")
        assert resp.status_code == 404

    def test_download_not_ready_returns_202(self, client):
        """Job that isn't done yet returns 202, not 404."""
        from app import _jobs, _jobs_lock
        import threading
        job_id = "test-pending-job"
        with _jobs_lock:
            _jobs[job_id] = {
                "status": "running",
                "file_path": None,
                "lock": threading.Lock(),
                "created": __import__("time").time(),
                "logs": [],
                "error": None,
                "progress": 50,
            }
        try:
            resp = client.get(f"/download/{job_id}")
            assert resp.status_code == 202
            # Job should still exist (not popped)
            with _jobs_lock:
                assert job_id in _jobs
        finally:
            with _jobs_lock:
                _jobs.pop(job_id, None)

    def test_download_done_job_returns_file(self, client, tmp_path):
        """Completed job can be downloaded, and multiple downloads work."""
        from app import _jobs, _jobs_lock
        import threading
        job_id = "test-done-job"
        # Create a temp file
        f = tmp_path / "report.xlsx"
        f.write_bytes(b"fake excel content")
        with _jobs_lock:
            _jobs[job_id] = {
                "status": "done",
                "file_path": str(f),
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "filename": "report.xlsx",
                "lock": threading.Lock(),
                "created": __import__("time").time(),
                "logs": [],
                "error": None,
                "progress": 100,
            }
        try:
            resp1 = client.get(f"/download/{job_id}")
            assert resp1.status_code == 200
            assert resp1.data == b"fake excel content"
            # Second download should also work (fix #2)
            resp2 = client.get(f"/download/{job_id}")
            assert resp2.status_code == 200
        finally:
            with _jobs_lock:
                _jobs.pop(job_id, None)

    def test_download_deleted_file_returns_410(self, client):
        """If temp file was already deleted, return 410 Gone."""
        from app import _jobs, _jobs_lock
        import threading
        job_id = "test-gone-job"
        with _jobs_lock:
            _jobs[job_id] = {
                "status": "done",
                "file_path": "/tmp/nonexistent-file-12345.xlsx",
                "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "filename": "report.xlsx",
                "lock": threading.Lock(),
                "created": __import__("time").time(),
                "logs": [],
                "error": None,
                "progress": 100,
            }
        try:
            resp = client.get(f"/download/{job_id}")
            assert resp.status_code == 410
        finally:
            with _jobs_lock:
                _jobs.pop(job_id, None)
