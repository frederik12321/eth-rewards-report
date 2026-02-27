#!/usr/bin/env python3
"""Ethereum staking tax report generator.

Uses Etherscan for withdrawal and block proposal history,
fetches ETH price data from multiple sources, and generates
Excel/CSV reports with monthly sheets + summary.
"""

import argparse
import bisect
import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font

# Ethereum beacon chain genesis timestamp (Dec 1, 2020 12:00:23 UTC)
GENESIS_TIMESTAMP = 1606824023
SECONDS_PER_SLOT = 12
GWEI_PER_ETH = Decimal(10**9)
WEI_PER_ETH = Decimal(10**18)

# Known MEV builder fee-recipient addresses (from ethstaker.tax, Feb 2026)
# Includes: Flashbots, beaverbuild, rsync-builder, Titan, bloXroute, Eden, etc.
KNOWN_BUILDER_ADDRESSES = {
    "0x001ee00bee25f81444e2d172773f37fe05ea2488",
    "0x089780a88f35b58144aa8a9be654207a1afe7959",
    "0x09fa51ab5387fb563200494e09e3c19cc0993c85",
    "0x1324c0fb6f45f3bf1aaa1fcdc08f17431f53ded7",
    "0x199d5ed7f45f4ee35960cf22eade2076e95b253f",
    "0x1d0124fee8dbe21884ab97adccbf5c55d768886e",
    "0x1f9090aae28b8a3dceadf281b0f12828e676c326",
    "0x229b8325bb9ac04602898b7e8989998710235d5f",
    "0x25d88437df70730122b73ef35462435d187c466f",
    "0x2cd54c2f60d94442ea38027df42663f1438e514b",
    "0x3b7faec3181114a99c243608bc822c5436441fff",
    "0x3c496df419762533607f30bb2143aff77bebc36a",
    "0x43dd22c94c1c1d46f4fcf664e2d7b11dee1d4154",
    "0x4460735849b78fd924cf0f21fca0ffc80c8b16cf",
    "0x473780deaf4a2ac070bbba936b0cdefe7f267dfc",
    "0x4838b106fce9647bdf1e7877bf73ce8b0bad5f97",
    "0x4a55474eacb48cefe25d7656db1976aa7ae70e3c",
    "0x57af10ed3469b2351ae60175d3c9b3740e1bb649",
    "0x5f927395213ee6b95de97bddcb1b2b1c0f16844f",
    "0x690b9a9e9aa1c9db991c7721a92d351db4fac990",
    "0x707fc1439cd11b34a984b989a18476c24a1182a1",
    "0x77777a6c097a1ce65c61a96a49bd1100f660ec94",
    "0x88c6c46ebf353a52bdbab708c23d0c81daa8134a",
    "0x8d5998a27b3cdf33479b65b18f075e20a7aa05b9",
    "0x95222290dd7278aa3ddd389cc1e1d165cc4bafe5",
    "0x9d8e2dc5615c674f329d18786d52af10a65af08b",
    "0xa7fdca7aa0b69927a34ec48ddcfe3d4c66ff0d94",
    "0xaab27b150451726ec7738aa1d0a94505c8729bd1",
    "0xae08c571e771f360c35f5715e36407ecc89d91ed",
    "0xb646d87963da1fb9d192ddba775f24f33e857128",
    "0xb64a30399f7f6b0c154c2e7af0a3ec7b0a5b131a",
    "0xbaf6dc2e647aeb6f510f9e318856a1bcd66c5e19",
    "0xbc178995898b0f611b4360df5ad653cdebe6de3f",
    "0xbd3afb0bb76683ecb4225f9dbc91f998713c3b01",
    "0xc08661e7d70f5a9f02e3e807e93cbef4747f861c",
    "0xc1612dc56c3e7e00d86c668df03904b7e59616c5",
    "0xc4b9beb1b7efb04deea31dc3b4c32a88ee210bf0",
    "0xd1a0b5843f384f92a6759015c742fc12d1d579a1",
    "0xdadb0d80178819f2319190d340ce9a924f783711",
    "0xdafea492d9c6733ae3d56b7ed1adb60692c98bc5",
    "0xdcca982701a264e8d629a6e8cfba9c1a27912623",
    "0xed7ce3de532213314bb07622d8bf606a4ba03cf1",
    "0xeeee755e55316154f4db6b9958f511a74e22e365",
    "0xeeee8db5fc7d505e99970945a9220ab7992050e3",
    "0xf2f5c73fa04406b1995e397b55c24ab1f3ea726c",
    "0xf573d99385c05c23b24ed33de616ad16a43a0919",
    "0xfee4446922f6e29834dea37d9e9192ccabf1e210",
    "0xfeebabe6b0418ec13b30aadf129f5dcdd4f70cea",
}


def slot_to_timestamp(slot: int) -> int:
    return GENESIS_TIMESTAMP + slot * SECONDS_PER_SLOT


def load_config(path: str = "config.yaml") -> dict:
    config_path = Path(path)
    if not config_path.exists():
        print(f"Error: {path} not found. Copy config.example.yaml and fill in your details.")
        sys.exit(1)
    with open(config_path) as f:
        return yaml.safe_load(f)


# ---------------------------------------------------------------------------
# Etherscan API client (primary data source for withdrawals + block rewards)
# ---------------------------------------------------------------------------

class EtherscanClient:
    BASE_URL = "https://api.etherscan.io/v2/api"

    def __init__(self, api_key: str, log_fn=None):
        self.api_key = api_key
        self.log = log_fn or print
        self._min_interval = 0.35  # max ~2.8 req/s — comfortable under 5/s free tier
        self._last_call = 0.0
        self.session = requests.Session()
        # Increase connection pool for sustained API usage
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=5,
            pool_maxsize=5,
            max_retries=0,  # we handle retries ourselves
        )
        self.session.mount("https://", adapter)

    def _throttle(self):
        """Ensure minimum interval between API calls to stay under rate limit."""
        elapsed = time.monotonic() - self._last_call
        if elapsed < self._min_interval:
            time.sleep(self._min_interval - elapsed)
        self._last_call = time.monotonic()

    def _get(self, params: Dict) -> Dict:
        params["chainid"] = 1
        params["apikey"] = self.api_key
        max_retries = 4
        for attempt in range(max_retries):
            self._throttle()
            try:
                resp = self.session.get(self.BASE_URL, params=params, timeout=30)
                resp.raise_for_status()
                data = resp.json()

                # Check for Etherscan rate-limit response
                if data.get("status") == "0" and "rate limit" in str(data.get("result", "")).lower():
                    if attempt < max_retries - 1:
                        wait = 2 ** (attempt + 1)
                        self.log(f"    Rate limited by Etherscan, waiting {wait}s...")
                        time.sleep(wait)
                        continue
                    else:
                        raise RuntimeError("Etherscan rate limit exceeded after retries")

                return data
            except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as e:
                if attempt < max_retries - 1:
                    wait = 2 ** (attempt + 1)
                    self.log(f"    Connection error, retrying in {wait}s... ({e})")
                    time.sleep(wait)
                else:
                    raise RuntimeError(f"Etherscan API connection failed after {max_retries} retries: {e}")

    def get_beacon_withdrawals(self, address: str, start_ts: int, end_ts: int) -> List[Dict]:
        """Fetch all beacon chain withdrawals for a withdrawal address."""
        self.log(f"  Fetching withdrawals from Etherscan for {address[:10]}...{address[-4:]}")
        all_withdrawals = []
        page = 1
        page_size = 10000

        while True:
            data = self._get({
                "module": "account",
                "action": "txsBeaconWithdrawal",
                "address": address,
                "startblock": 0,
                "endblock": 99999999,
                "page": page,
                "offset": page_size,
                "sort": "asc",
            })

            if data.get("status") != "1":
                result_msg = data.get("result", "")
                message = data.get("message", "")
                if "No transactions found" in str(message) or "No transactions found" in str(result_msg):
                    self.log(f"  No beacon withdrawals found for this address")
                    break
                if "Invalid API Key" in str(result_msg) or "Missing" in str(result_msg):
                    raise RuntimeError(f"Etherscan API key error: {result_msg}")
                if page == 1:
                    self.log(f"  Etherscan withdrawals response: status={message}, result={result_msg}")
                break

            results = data.get("result", [])
            if not results:
                break

            for w in results:
                ts = int(w["timestamp"])
                if ts < start_ts or ts > end_ts:
                    continue
                amount_gwei = int(w["amount"])
                if amount_gwei < 0:
                    continue  # Skip invalid negative amounts
                all_withdrawals.append({
                    "timestamp": ts,
                    "validator": int(w["validatorIndex"]),
                    "amount_gwei": amount_gwei,
                    "amount_eth": Decimal(w["amount"]) / GWEI_PER_ETH,
                    "block_number": int(w["blockNumber"]),
                    "type": "withdrawal",
                })

            if len(results) < page_size:
                break
            page += 1

        self.log(f"  Fetched {len(all_withdrawals)} withdrawals in date range")
        return all_withdrawals

    def _get_block_fee_recipient(self, block_number: int) -> str:
        """Get the feeRecipient (miner) of a specific block."""
        data = self._get({
            "module": "proxy",
            "action": "eth_getBlockByNumber",
            "tag": hex(block_number),
            "boolean": "false",
        })
        result = data.get("result", {})
        if not result:
            return ""
        return result.get("miner", "").lower()

    def _fetch_paginated(self, params_base: Dict) -> List[Dict]:
        """Fetch all pages from an Etherscan list endpoint."""
        all_results = []
        page = 1
        while True:
            params = {**params_base, "page": page, "offset": 10000}
            data = self._get(params)
            if data.get("status") != "1":
                break
            results = data.get("result", [])
            if not results:
                break
            all_results.extend(results)
            if len(results) < 10000:
                break
            page += 1
        return all_results

    def _get_block_number_by_timestamp(self, timestamp: int, closest: str = "before") -> int:
        """Convert a timestamp to the nearest block number via Etherscan API."""
        data = self._get({
            "module": "block",
            "action": "getblocknobytime",
            "timestamp": str(timestamp),
            "closest": closest,
        })
        result = data.get("result", "")
        return int(result) if result and result.isdigit() else 0

    def get_execution_rewards(self, address: str, start_ts: int, end_ts: int) -> List[Dict]:
        """Fetch all execution layer rewards: local blocks (priority fees) + MEV payments.

        Detection methods:
        1. getminedblocks — local blocks where user is feeRecipient
        2. txlist + txlistinternal — incoming transactions checked against:
           a. Known builder address list (fast, no extra API call)
           b. Block feeRecipient lookup (fallback for unknown builders)
        """
        address_lower = address.lower()

        # Convert timestamps to block numbers to narrow transaction queries
        start_block = self._get_block_number_by_timestamp(start_ts, "after") or 0
        end_block = self._get_block_number_by_timestamp(end_ts, "before") or 99999999

        # --- Step 1: Local blocks (non-MEV, user is feeRecipient) ---
        self.log(f"  Fetching local block rewards from Etherscan...")
        local_blocks = {}
        results = self._fetch_paginated({
            "module": "account",
            "action": "getminedblocks",
            "address": address,
            "blocktype": "blocks",
        })
        for b in results:
            ts = int(b["timeStamp"])
            if start_ts <= ts <= end_ts:
                bn = int(b["blockNumber"])
                local_blocks[bn] = {
                    "timestamp": ts,
                    "block_number": bn,
                    "amount_eth": Decimal(b.get("blockReward", "0")) / WEI_PER_ETH,
                    "type": "block_reward",
                }
        self.log(f"    Found {len(local_blocks)} local block rewards")

        # --- Step 2: Find MEV candidates from regular + internal transactions ---
        self.log(f"  Fetching incoming transactions to detect MEV payments...")
        candidate_txs = []  # list of (block_number, sender, value_wei, timestamp)

        # 2a: Regular transactions (txlist) — filtered by block range
        regular_txs = self._fetch_paginated({
            "module": "account",
            "action": "txlist",
            "address": address,
            "startblock": start_block,
            "endblock": end_block,
            "sort": "asc",
        })
        for tx in regular_txs:
            if tx.get("to", "").lower() != address_lower:
                continue
            value = int(tx.get("value", 0))
            if value == 0:
                continue
            ts = int(tx["timeStamp"])
            if start_ts <= ts <= end_ts:
                bn = int(tx["blockNumber"])
                if bn not in local_blocks:
                    candidate_txs.append((bn, tx["from"].lower(), str(value), ts))

        # 2b: Internal transactions (txlistinternal) — catches MEV via smart contracts
        self.log(f"  Fetching internal transactions to detect MEV via contracts...")
        internal_txs = self._fetch_paginated({
            "module": "account",
            "action": "txlistinternal",
            "address": address,
            "startblock": start_block,
            "endblock": end_block,
            "sort": "asc",
        })
        seen_blocks = set(c[0] for c in candidate_txs)
        for tx in internal_txs:
            if tx.get("to", "").lower() != address_lower:
                continue
            value = int(tx.get("value", 0))
            if value == 0:
                continue
            ts = int(tx["timeStamp"])
            if start_ts <= ts <= end_ts:
                bn = int(tx["blockNumber"])
                if bn not in local_blocks and bn not in seen_blocks:
                    candidate_txs.append((bn, tx["from"].lower(), str(value), ts))
                    seen_blocks.add(bn)

        self.log(f"    Found {len(candidate_txs)} candidate transactions (regular + internal)")

        # --- Step 3: Classify candidates as MEV ---
        mev_rewards = {}
        block_cache = {}
        builder_hits = 0
        lookup_hits = 0

        for i, (bn, sender, value_str, ts) in enumerate(candidate_txs):
            is_mev = False

            # Fast path: sender is a known builder address
            if sender in KNOWN_BUILDER_ADDRESSES:
                is_mev = True
                builder_hits += 1
            else:
                # Slow path: look up block's feeRecipient
                if bn not in block_cache:
                    block_cache[bn] = self._get_block_fee_recipient(bn)
                fee_recipient = block_cache[bn]
                if fee_recipient and sender == fee_recipient:
                    is_mev = True
                    lookup_hits += 1

            if is_mev:
                mev_rewards[bn] = {
                    "timestamp": ts,
                    "block_number": bn,
                    "amount_eth": Decimal(value_str) / WEI_PER_ETH,
                    "type": "block_reward",
                }

            if (i + 1) % 10 == 0 or (i + 1) == len(candidate_txs):
                self.log(f"    Checked {i + 1}/{len(candidate_txs)} transactions ({len(mev_rewards)} MEV found so far)")

        self.log(f"    Found {len(mev_rewards)} MEV block rewards ({builder_hits} via known builders, {lookup_hits} via block lookup)")

        # --- Step 4: Combine local + MEV ---
        all_rewards = list(local_blocks.values()) + list(mev_rewards.values())
        all_rewards.sort(key=lambda x: x["timestamp"])
        self.log(f"  Total execution rewards: {len(all_rewards)} ({len(local_blocks)} local + {len(mev_rewards)} MEV)")
        return all_rewards

    # EIP-7002: Execution layer withdrawal request contract (Pectra)
    WITHDRAWAL_REQUEST_CONTRACT = "0x00000961Ef480Eb55e80D19ad83579A64c007002".lower()

    def get_withdrawal_requests(self, address: str, start_ts: int, end_ts: int) -> Dict[str, List[Dict]]:
        """Fetch EIP-7002 withdrawal requests sent from this address.

        Returns dict mapping validator pubkey → list of requests [{timestamp, amount_gwei}].
        Amount=0 means full exit request.
        """
        self.log(f"  Fetching EIP-7002 withdrawal requests for {address[:10]}...{address[-4:]}")
        results = self._fetch_paginated({
            "module": "account",
            "action": "txlist",
            "address": address,
            "startblock": 0,
            "endblock": 99999999,
            "sort": "asc",
        })

        requests_by_pubkey: Dict[str, List[Dict]] = {}
        for tx in results:
            if tx.get("to", "").lower() != self.WITHDRAWAL_REQUEST_CONTRACT:
                continue
            if tx.get("from", "").lower() != address.lower():
                continue
            ts = int(tx["timeStamp"])
            if ts < start_ts or ts > end_ts:
                continue
            if tx.get("isError", "0") == "1":
                continue

            input_data = tx.get("input", "")
            # Input: 20B sender + 48B validator pubkey + 8B amount (gwei)
            # In tx.input the 20B sender is sometimes omitted; raw data starts at pubkey
            # Standard format: 0x + 56 bytes hex (112 chars) for sender-included
            # or 0x + 112 hex chars for full 56 bytes
            hex_data = input_data[2:] if input_data.startswith("0x") else input_data
            if len(hex_data) >= 112:
                # Skip 20-byte sender prefix, extract 48-byte pubkey + 8-byte amount
                pubkey = "0x" + hex_data[40:136]
                amount_hex = hex_data[136:152] if len(hex_data) >= 152 else "0"
                amount_gwei = int(amount_hex, 16) if amount_hex else 0
            elif len(hex_data) >= 96:
                # No sender prefix: 48-byte pubkey + 8-byte amount
                pubkey = "0x" + hex_data[:96]
                amount_hex = hex_data[96:112] if len(hex_data) >= 112 else "0"
                amount_gwei = int(amount_hex, 16) if amount_hex else 0
            else:
                continue  # unexpected format

            pubkey_lower = pubkey.lower()
            if pubkey_lower not in requests_by_pubkey:
                requests_by_pubkey[pubkey_lower] = []
            requests_by_pubkey[pubkey_lower].append({
                "timestamp": ts,
                "amount_gwei": amount_gwei,
                "full_exit": amount_gwei == 0,
            })

        total = sum(len(v) for v in requests_by_pubkey.values())
        self.log(f"    Found {total} withdrawal requests for {len(requests_by_pubkey)} validator(s)")
        return requests_by_pubkey

    # EIP-7251: Consolidation request contract (Pectra)
    CONSOLIDATION_REQUEST_CONTRACT = "0x0000BBdDc7CE488642fb579F8B00f3a590007251".lower()

    def get_consolidation_events(self, address: str, start_ts: int, end_ts: int) -> List[Dict]:
        """Fetch EIP-7251 consolidation requests sent from this address.

        Returns list of {source_pubkey, target_pubkey, timestamp, is_upgrade}.
        is_upgrade=True when source==target (0x01→0x02 credential conversion).
        """
        self.log(f"  Fetching EIP-7251 consolidation events for {address[:10]}...{address[-4:]}")
        results = self._fetch_paginated({
            "module": "account",
            "action": "txlist",
            "address": address,
            "startblock": 0,
            "endblock": 99999999,
            "sort": "asc",
        })

        consolidations = []
        for tx in results:
            if tx.get("to", "").lower() != self.CONSOLIDATION_REQUEST_CONTRACT:
                continue
            if tx.get("from", "").lower() != address.lower():
                continue
            ts = int(tx["timeStamp"])
            if ts < start_ts or ts > end_ts:
                continue
            if tx.get("isError", "0") == "1":
                continue

            input_data = tx.get("input", "")
            hex_data = input_data[2:] if input_data.startswith("0x") else input_data
            # Input: 20B sender + 48B source pubkey + 48B target pubkey
            if len(hex_data) >= 192:
                # With 20-byte sender prefix
                source_pubkey = "0x" + hex_data[40:136]
                target_pubkey = "0x" + hex_data[136:232]
            elif len(hex_data) >= 192 - 40:
                # Without sender prefix: 48B source + 48B target
                source_pubkey = "0x" + hex_data[:96]
                target_pubkey = "0x" + hex_data[96:192]
            else:
                continue

            consolidations.append({
                "source_pubkey": source_pubkey.lower(),
                "target_pubkey": target_pubkey.lower(),
                "timestamp": ts,
                "is_upgrade": source_pubkey.lower() == target_pubkey.lower(),
            })

        upgrades = sum(1 for c in consolidations if c["is_upgrade"])
        merges = len(consolidations) - upgrades
        self.log(f"    Found {len(consolidations)} consolidation events ({upgrades} upgrades, {merges} merges)")
        return consolidations


# ---------------------------------------------------------------------------
# Beacon chain API client (resolves validator index/pubkey → withdrawal address)
# ---------------------------------------------------------------------------

class BeaconClient:
    """Resolves validator index or public key to withdrawal address via standard Beacon API."""

    DEFAULT_URL = "https://ethereum-beacon-api.publicnode.com"

    def __init__(self, base_url: Optional[str] = None, log_fn=None):
        self.base_url = (base_url or self.DEFAULT_URL).rstrip("/")
        self.log = log_fn or print
        self.session = requests.Session()

    def get_validator_info(self, validator_id: str) -> Dict:
        """Look up validator by index or public key.

        Returns dict with 'withdrawal_address' and 'validator_index'.
        Raises RuntimeError if not found or has BLS (0x00) credentials.
        """
        url = f"{self.base_url}/eth/v1/beacon/states/head/validators/{validator_id}"
        max_retries = 3
        resp = None
        for attempt in range(max_retries):
            try:
                resp = self.session.get(url, timeout=30)
                break
            except requests.exceptions.RequestException as e:
                if attempt < max_retries - 1:
                    wait = 2 ** (attempt + 1)
                    self.log(f"    Beacon API request failed, retrying in {wait}s... ({e})")
                    time.sleep(wait)
                else:
                    raise RuntimeError(f"Beacon API request failed after {max_retries} retries: {e}")

        if resp.status_code == 404:
            raise RuntimeError(f"Validator '{validator_id}' not found on the beacon chain.")
        resp.raise_for_status()

        data = resp.json().get("data", {})
        index = int(data.get("index", 0))
        creds = data.get("validator", {}).get("withdrawal_credentials", "")

        if creds.startswith("0x00"):
            raise RuntimeError(
                f"Validator {validator_id} uses BLS withdrawal credentials (0x00). "
                "It has not set an execution layer withdrawal address yet."
            )

        # 0x01 or 0x02 credentials: last 20 bytes = withdrawal address
        credential_type = creds[:4]  # "0x01" or "0x02"
        address = "0x" + creds[-40:]
        pubkey = data.get("validator", {}).get("pubkey", "")
        balance_gwei = int(data.get("balance", 0))
        effective_balance_gwei = int(data.get("validator", {}).get("effective_balance", 0))
        status = data.get("status", "unknown")
        balance_eth = Decimal(balance_gwei) / GWEI_PER_ETH
        effective_balance_eth = Decimal(effective_balance_gwei) / GWEI_PER_ETH
        cred_label = "compounding" if credential_type == "0x02" else "distributing"
        self.log(f"  Resolved validator {validator_id} → index {index}, {cred_label} ({credential_type}), "
              f"balance {balance_eth:.4f} ETH, address {address[:10]}...{address[-4:]}")
        return {
            "withdrawal_address": address,
            "validator_index": index,
            "credential_type": credential_type,
            "pubkey": pubkey,
            "balance_eth": balance_eth,
            "effective_balance_eth": effective_balance_eth,
            "status": status,
        }


def parse_validator_input(value: str) -> Dict:
    """Parse user input and detect whether it's an address, validator index(es), or pubkey.

    Supports comma-separated validator indices (e.g. "652648, 652649, 652650").
    Returns dict with 'type' and 'value' keys.
    For multiple indices, type is 'validator_indices' and value is a list of strings.
    """
    value = value.strip()
    if not value:
        raise ValueError("Empty input")

    # Check for comma-separated values (multiple validator indices)
    if "," in value:
        parts = [p.strip() for p in value.split(",") if p.strip()]
        for p in parts:
            if not p.isdigit():
                raise ValueError(
                    f"Invalid validator index '{p}' in comma-separated list. "
                    "All values must be numeric validator indices."
                )
        return {"type": "validator_indices", "value": parts}

    if value.isdigit():
        return {"type": "validator_index", "value": value}

    if value.startswith("0x") or value.startswith("0X"):
        hex_part = value[2:]
        if len(hex_part) == 40 and all(c in "0123456789abcdefABCDEF" for c in hex_part):
            return {"type": "withdrawal_address", "value": value}
        if len(hex_part) == 96 and all(c in "0123456789abcdefABCDEF" for c in hex_part):
            return {"type": "validator_pubkey", "value": value}

    # 96 hex chars without 0x prefix
    if len(value) == 96 and all(c in "0123456789abcdefABCDEF" for c in value):
        return {"type": "validator_pubkey", "value": "0x" + value}

    raise ValueError(
        "Unrecognized input format. Expected: "
        "withdrawal address (0x + 40 hex chars), "
        "validator index (number or comma-separated numbers), "
        "or public key (0x + 96 hex chars)"
    )


# ---------------------------------------------------------------------------
# Price providers
# ---------------------------------------------------------------------------

SUPPORTED_CURRENCIES = ["EUR", "USD", "GBP", "CHF", "CAD", "AUD", "JPY", "SEK", "NOK", "DKK"]


class PriceCache:
    """SQLite cache for historical ETH prices. Prices never change, so cache forever."""

    def __init__(self, db_path: str = "price_cache.db"):
        import sqlite3
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, timeout=10)
        self.conn.execute("PRAGMA journal_mode=WAL")
        self.conn.execute("PRAGMA busy_timeout=5000")
        self.conn.execute(
            "CREATE TABLE IF NOT EXISTS prices "
            "(timestamp INTEGER, currency TEXT, price TEXT, "
            "PRIMARY KEY (timestamp, currency))"
        )
        self.conn.commit()

    def close(self):
        """Explicitly close the database connection."""
        if self.conn:
            self.conn.close()
            self.conn = None

    def get_cached(self, currency: str, start_ts: int, end_ts: int) -> List[Tuple[int, Decimal]]:
        rows = self.conn.execute(
            "SELECT timestamp, price FROM prices "
            "WHERE currency = ? AND timestamp >= ? AND timestamp <= ? "
            "ORDER BY timestamp",
            (currency, start_ts, end_ts),
        ).fetchall()
        return [(r[0], Decimal(r[1])) for r in rows]

    def store(self, currency: str, prices: List[Tuple[int, Decimal]]):
        if not prices:
            return
        self.conn.executemany(
            "INSERT OR IGNORE INTO prices (timestamp, currency, price) VALUES (?, ?, ?)",
            [(ts, currency, str(price)) for ts, price in prices],
        )
        self.conn.commit()

    def find_gaps(self, currency: str, start_ts: int, end_ts: int) -> List[Tuple[int, int]]:
        """Find time ranges not yet cached (at hourly resolution).

        Returns list of (gap_start, gap_end) tuples that need fetching.
        """
        cached = self.conn.execute(
            "SELECT MIN(timestamp), MAX(timestamp) FROM prices "
            "WHERE currency = ? AND timestamp >= ? AND timestamp <= ?",
            (currency, start_ts, end_ts),
        ).fetchone()

        if cached[0] is None:
            # Nothing cached for this range
            return [(start_ts, end_ts)]

        gaps = []
        cached_min, cached_max = cached

        # Gap before cached data
        if start_ts < cached_min - 3600:
            gaps.append((start_ts, cached_min))
        # Gap after cached data
        if end_ts > cached_max + 3600:
            gaps.append((cached_max, end_ts))

        return gaps


class CryptoCompareClient:
    BASE_URL = "https://min-api.cryptocompare.com/data/v2"

    def __init__(self, currency: str = "EUR", cache_path: str = "price_cache.db", log_fn=None, api_key: str = "", coingecko_api_key: str = "", verbose_sources: bool = False):
        self.currency = currency.upper()
        self.log = log_fn or print
        self._verbose = verbose_sources
        self.session = requests.Session()
        if api_key:
            self.session.headers["authorization"] = f"Apikey {api_key}"
        self.coingecko_api_key = coingecko_api_key
        self.cache = PriceCache(cache_path)

    def close(self):
        """Close the price cache database connection."""
        self.cache.close()

    def _fetch_forex_rates(self, start_date: str, end_date: str) -> Dict[str, Decimal]:
        """Fetch daily USD→target forex rates from Frankfurter (ECB data, no key needed).

        Returns dict mapping 'YYYY-MM-DD' -> conversion rate.
        """
        if self.currency == "USD":
            return {}
        try:
            resp = self.session.get(
                f"https://api.frankfurter.dev/v1/{start_date}..{end_date}",
                params={"base": "USD", "symbols": self.currency},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            rates = {}
            for date_str, rate_dict in data.get("rates", {}).items():
                if self.currency in rate_dict:
                    rates[date_str] = Decimal(str(rate_dict[self.currency]))
            return rates
        except Exception as e:
            if self._verbose:
                self.log(f"    Frankfurter forex API failed: {e}")
            return {}

    def _fetch_range_defillama(self, start_ts: int, end_ts: int) -> List[Tuple[int, Decimal]]:
        """Fallback: fetch ETH/USD from DeFiLlama and convert via ECB forex rates.

        No API key needed. DeFiLlama returns hourly data, Frankfurter provides
        daily USD→target conversion rates.
        """
        from datetime import datetime as _dt

        # Step 1: Fetch forex rates if non-USD
        forex_rates = {}
        if self.currency != "USD":
            start_date = _dt.utcfromtimestamp(start_ts).strftime("%Y-%m-%d")
            end_date = _dt.utcfromtimestamp(end_ts).strftime("%Y-%m-%d")
            if self._verbose:
                self.log(f"    Fetching USD/{self.currency} forex rates from ECB...")
            forex_rates = self._fetch_forex_rates(start_date, end_date)
            if not forex_rates:
                self.log(f"    Could not fetch conversion rates for {self.currency}")
                return []
            if self._verbose:
                self.log(f"    Got {len(forex_rates)} daily forex rates")

        # Step 2: Fetch ETH/USD from DeFiLlama in chunks of 500 hours (API limit)
        fetched = []
        chunk_hours = 500
        current_start = start_ts
        call_count = 0
        max_retries = 3

        while current_start < end_ts:
            remaining_hours = min(chunk_hours, max(1, int((end_ts - current_start) / 3600)))

            data = None
            for attempt in range(max_retries):
                try:
                    resp = self.session.get(
                        "https://coins.llama.fi/chart/coingecko:ethereum",
                        params={
                            "start": current_start,
                            "span": remaining_hours,
                            "period": "1h",
                            "searchWidth": 600,
                        },
                        timeout=30,
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    break
                except (requests.exceptions.ConnectionError,
                        requests.exceptions.Timeout,
                        requests.exceptions.HTTPError) as e:
                    if attempt < max_retries - 1:
                        wait = 2 ** (attempt + 1)
                        self.log(f"    Price request failed, retrying in {wait}s...")
                        time.sleep(wait)
                    else:
                        self.log(f"    Price fetch failed after {max_retries} retries")
                        return fetched

            if data is None:
                break

            prices_data = []
            coins = data.get("coins", {})
            for coin_data in coins.values():
                prices_data = coin_data.get("prices", [])

            if not prices_data:
                self.log(f"    No price data returned")
                break

            for entry in prices_data:
                ts = entry["timestamp"]
                usd_price = Decimal(str(entry["price"]))
                if start_ts <= ts <= end_ts and usd_price > 0:
                    if self.currency == "USD":
                        fetched.append((ts, usd_price))
                    else:
                        # Convert USD→target using nearest daily forex rate
                        day_str = _dt.utcfromtimestamp(ts).strftime("%Y-%m-%d")
                        rate = forex_rates.get(day_str)
                        if rate is None:
                            # Find nearest available rate
                            sorted_dates = sorted(forex_rates.keys())
                            if sorted_dates:
                                nearest = min(sorted_dates, key=lambda d: abs(
                                    _dt.strptime(d, "%Y-%m-%d").timestamp() - ts))
                                rate = forex_rates[nearest]
                        if rate:
                            fetched.append((ts, usd_price * rate))

            call_count += 1
            if call_count % 5 == 0:
                self.log(f"    Fetched {len(fetched)} price points so far ({call_count} API calls)...")

            current_start = current_start + remaining_hours * 3600
            time.sleep(0.3)

        return fetched

    def _fetch_range_coingecko(self, start_ts: int, end_ts: int) -> List[Tuple[int, Decimal]]:
        """Fallback: fetch prices from CoinGecko for currencies CryptoCompare doesn't support.

        CoinGecko free tier returns hourly data for ranges <=90 days,
        daily data for ranges >90 days. We chunk into 90-day windows.
        """
        fetched = []
        chunk_seconds = 89 * 86400  # 89 days per chunk to stay under 90-day hourly limit
        current_start = start_ts
        call_count = 0
        max_retries = 3

        while current_start < end_ts:
            chunk_end = min(current_start + chunk_seconds, end_ts)

            data = None
            for attempt in range(max_retries):
                try:
                    headers = {}
                    if self.coingecko_api_key:
                        headers["x-cg-demo-api-key"] = self.coingecko_api_key
                    resp = self.session.get(
                        "https://api.coingecko.com/api/v3/coins/ethereum/market_chart/range",
                        params={
                            "vs_currency": self.currency.lower(),
                            "from": current_start,
                            "to": chunk_end,
                        },
                        headers=headers,
                        timeout=30,
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    break
                except (requests.exceptions.ConnectionError,
                        requests.exceptions.Timeout,
                        requests.exceptions.HTTPError) as e:
                    if attempt < max_retries - 1:
                        wait = 2 ** (attempt + 1)
                        msg = f"    CoinGecko request failed, retrying in {wait}s... ({e})" if self._verbose else f"    Price request failed, retrying in {wait}s..."
                        self.log(msg)
                        time.sleep(wait)
                    else:
                        msg = f"    CoinGecko failed after {max_retries} retries: {e}" if self._verbose else f"    Price fetch failed after {max_retries} retries"
                        self.log(msg)
                        return fetched

            if data is None or "prices" not in data:
                self.log(f"    No price data returned")
                break

            for ts_ms, price_val in data["prices"]:
                ts = int(ts_ms / 1000)
                price = Decimal(str(price_val))
                if start_ts <= ts <= end_ts and price > 0:
                    fetched.append((ts, price))

            call_count += 1
            if call_count % 5 == 0:
                self.log(f"    Fetched {len(fetched)} price points so far ({call_count} API calls)...")

            current_start = chunk_end + 1
            time.sleep(1.5)  # CoinGecko free tier: ~30 req/min

        return fetched

    def _fetch_range(self, start_ts: int, end_ts: int) -> List[Tuple[int, Decimal]]:
        """Fetch hourly prices from CryptoCompare for a specific time range."""
        fetched = []
        current_to = end_ts
        call_count = 0
        max_retries = 3

        while current_to > start_ts:
            params = {
                "fsym": "ETH",
                "tsym": self.currency,
                "limit": 2000,
                "toTs": current_to,
            }

            # Retry loop for each API call
            data = None
            for attempt in range(max_retries):
                try:
                    resp = self.session.get(
                        f"{self.BASE_URL}/histohour",
                        params=params,
                        timeout=30,
                    )
                    resp.raise_for_status()
                    data = resp.json()
                    break
                except (requests.exceptions.ConnectionError,
                        requests.exceptions.Timeout,
                        requests.exceptions.HTTPError) as e:
                    if attempt < max_retries - 1:
                        wait = 2 ** (attempt + 1)
                        msg = f"    CryptoCompare request failed, retrying in {wait}s... ({e})" if self._verbose else f"    Price request failed, retrying in {wait}s..."
                        self.log(msg)
                        time.sleep(wait)
                    else:
                        msg = f"    CryptoCompare failed after {max_retries} retries: {e}" if self._verbose else f"    Price fetch failed after {max_retries} retries"
                        self.log(msg)
                        return fetched  # Return what we have so far

            if data is None or data.get("Response") != "Success":
                if data and self._verbose:
                    self.log(f"    CryptoCompare error: {data.get('Message')}")
                break

            entries = data["Data"]["Data"]
            for entry in entries:
                ts = entry["time"]
                price = Decimal(str(entry["close"]))
                if start_ts <= ts <= end_ts and price > 0:
                    fetched.append((ts, price))

            if entries:
                earliest = entries[0]["time"]
                if earliest >= current_to:
                    break
                current_to = earliest
            else:
                break

            call_count += 1
            if call_count % 5 == 0:
                self.log(f"    Fetched {len(fetched)} price points so far...")
            time.sleep(0.3)

        return fetched

    def get_prices(self, start_ts: int, end_ts: int) -> List[Tuple[int, Decimal]]:
        """Fetch hourly ETH prices, using local cache where possible.

        First checks SQLite cache for existing data, then fetches only
        the missing time ranges from CryptoCompare. Newly fetched prices
        are stored in the cache for future use.
        """
        # Check what we already have cached
        cached = self.cache.get_cached(self.currency, start_ts, end_ts)
        gaps = self.cache.find_gaps(self.currency, start_ts, end_ts)

        if not gaps:
            self.log(f"  Loaded {len(cached)} ETH/{self.currency} price points (hourly)")
            return cached

        if cached:
            self.log(f"  Loaded {len(cached)} price points, fetching {len(gaps)} gap(s)...")
        else:
            self.log(f"  Fetching ETH/{self.currency} prices...")

        # Fetch missing ranges
        all_fetched = []
        for gap_start, gap_end in gaps:
            fetched = self._fetch_range(gap_start, gap_end)
            if not fetched:
                # Primary source has no data — try alternative sources
                if self._verbose:
                    self.log(f"  Primary source has no data for ETH/{self.currency}, trying alternatives...")
                fetched = self._fetch_range_defillama(gap_start, gap_end)
            if not fetched and self.coingecko_api_key:
                fetched = self._fetch_range_coingecko(gap_start, gap_end)
            all_fetched.extend(fetched)

        # Store newly fetched prices in cache
        if all_fetched:
            self.cache.store(self.currency, all_fetched)
            if self._verbose:
                self.log(f"  Cached {len(all_fetched)} new price points")

        # Combine cached + fetched, deduplicate, sort
        combined = cached + all_fetched
        combined.sort(key=lambda x: x[0])
        seen = set()
        deduped = []
        for ts, price in combined:
            if ts not in seen:
                seen.add(ts)
                deduped.append((ts, price))

        # Coverage check
        expected_hours = max(1, (end_ts - start_ts) / 3600)
        coverage = len(deduped) / expected_hours
        if coverage < 0.8:
            self.log(f"  WARNING: Only {coverage:.0%} price coverage ({len(deduped)}/{int(expected_hours)} hours). Some events may use approximate prices.")

        self.log(f"  Total: {len(deduped)} price data points ({self.currency}, hourly)")
        return deduped


def find_nearest_price(prices: List[Tuple[int, Decimal]], timestamp: int) -> Decimal:
    """Find the price closest to the given timestamp using binary search."""
    if not prices:
        return Decimal(0)
    timestamps = [p[0] for p in prices]
    idx = bisect.bisect_left(timestamps, timestamp)
    if idx == 0:
        return prices[0][1]
    if idx >= len(prices):
        return prices[-1][1]
    # Pick the closer of the two neighbors
    before = prices[idx - 1]
    after = prices[idx]
    if abs(timestamp - before[0]) <= abs(timestamp - after[0]):
        return before[1]
    return after[1]


# ---------------------------------------------------------------------------
# Report writers (file + in-memory)
# ---------------------------------------------------------------------------

def _prepare_events(events: List[Dict], prices: List[Tuple[int, Decimal]], tz=None):
    """Enrich events with price data and datetime. All amounts use Decimal.

    tz: optional timezone for datetime display (default: UTC).
    """
    display_tz = tz or timezone.utc
    for event in sorted(events, key=lambda e: e["timestamp"]):
        dt = datetime.fromtimestamp(event["timestamp"], tz=timezone.utc).astimezone(display_tz)
        price = find_nearest_price(prices, event["timestamp"])
        event["price_fiat"] = price
        event["value_fiat"] = event["amount_eth"] * price
        event["datetime"] = dt


def _write_csv_rows(events: List[Dict], currency: str, writer, tz_label: str = "UTC"):
    """Write CSV rows to any csv.writer-compatible object."""
    writer.writerow(["Date", f"Time ({tz_label})", "Type", "Validator", "Amount [ETH]", f"Price [{currency}/ETH]", f"Value [{currency}]"])
    for event in sorted(events, key=lambda e: e["timestamp"]):
        dt = event["datetime"]
        writer.writerow([
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M:%S"),
            event["type"],
            event.get("validator", ""),
            f"{event['amount_eth']:.9f}",
            f"{event['price_fiat']:.2f}",
            f"{event['value_fiat']:.2f}",
        ])


def write_csv(events: List[Dict], prices: List[Tuple[int, Decimal]], account_name: str, date_from: str, date_to: str, currency: str = "EUR", output_dir: str = "."):
    """Write a flat CSV file with all events to disk."""
    import csv
    _prepare_events(events, prices)

    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "", account_name) or "unnamed"
    filename = f"staking_report_{safe_name}_{date_from}_{date_to}.csv"
    filepath = os.path.join(output_dir, filename)

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        _write_csv_rows(events, currency, writer)

    print(f"\n  Saved CSV: {filepath}")
    return filepath


def write_csv_bytes(events: List[Dict], prices: List[Tuple[int, Decimal]], currency: str = "EUR", tz_label: str = "UTC") -> bytes:
    """Generate CSV report in memory and return as bytes."""
    import csv
    import io
    _prepare_events(events, prices)

    buf = io.StringIO()
    writer = csv.writer(buf)
    _write_csv_rows(events, currency, writer, tz_label=tz_label)
    return buf.getvalue().encode("utf-8")


def _build_excel_workbook(events: List[Dict], prices: List[Tuple[int, Decimal]], currency: str = "EUR",
                          validator_info: Optional[List[Dict]] = None, tz_label: str = "UTC") -> Workbook:
    """Build an Excel workbook with monthly sheets + summary. Returns the Workbook object.

    validator_info: optional list of dicts with keys: index, name, credential_type, status,
                    balance_eth, effective_balance_eth, withdrawn_eth, withdrawn_income_eth.
    """
    _prepare_events(events, prices)
    wb = Workbook()

    # Group events by month
    events_by_month = defaultdict(list)
    for event in sorted(events, key=lambda e: e["timestamp"]):
        month_key = event["datetime"].strftime("%Y-%m")
        events_by_month[month_key].append(event)

    # Create monthly sheets
    header = ["Date", f"Time ({tz_label})", "Type", "Validator", "Amount [ETH]", f"Price [{currency}/ETH]", f"Value [{currency}]"]
    header_font = Font(bold=True)

    sheet_names = sorted(events_by_month.keys())

    # Remove default sheet
    wb.remove(wb.active)

    for month_key in sheet_names:
        ws = wb.create_sheet(title=month_key)
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font

        for event in events_by_month[month_key]:
            dt = event["datetime"]
            validator = event.get("validator", "")
            ws.append([
                dt.strftime("%Y-%m-%d"),
                dt.strftime("%H:%M:%S"),
                event["type"],
                validator,
                event["amount_eth"],
                event["price_fiat"],
                event["value_fiat"],
            ])

        # Format number columns
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = "0.000000000"
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = "#,##0.00"
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = "#,##0.00"

        # Auto-size columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Check if any principal withdrawals exist (0x02 validators)
    has_principal = any(e["type"] == "withdrawal_principal" for e in events)

    # Summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    summary_header = [
        "Month",
        "Withdrawals [ETH]", f"Withdrawals [{currency}]",
        "Block Rewards [ETH]", f"Block Rewards [{currency}]",
        "Total Income [ETH]", f"Total Income [{currency}]",
    ]
    if has_principal:
        summary_header.extend([f"Principal [ETH]", f"Principal [{currency}]"])
    ws_summary.append(summary_header)
    for cell in ws_summary[1]:
        cell.font = header_font

    grand_totals = {
        "w_eth": Decimal(0), "w_fiat": Decimal(0),
        "br_eth": Decimal(0), "br_fiat": Decimal(0),
        "total_eth": Decimal(0), "total_fiat": Decimal(0),
        "p_eth": Decimal(0), "p_fiat": Decimal(0),
    }

    for month_key in sheet_names:
        month_events = events_by_month[month_key]
        w_eth = sum(e["amount_eth"] for e in month_events if e["type"] == "withdrawal")
        w_fiat = sum(e["value_fiat"] for e in month_events if e["type"] == "withdrawal")
        br_eth = sum(e["amount_eth"] for e in month_events if e["type"] == "block_reward")
        br_fiat = sum(e["value_fiat"] for e in month_events if e["type"] == "block_reward")
        p_eth = sum(e["amount_eth"] for e in month_events if e["type"] == "withdrawal_principal")
        p_fiat = sum(e["value_fiat"] for e in month_events if e["type"] == "withdrawal_principal")
        total_eth = w_eth + br_eth
        total_fiat = w_fiat + br_fiat

        row_data = [month_key, w_eth, w_fiat, br_eth, br_fiat, total_eth, total_fiat]
        if has_principal:
            row_data.extend([p_eth, p_fiat])
        ws_summary.append(row_data)

        grand_totals["w_eth"] += w_eth
        grand_totals["w_fiat"] += w_fiat
        grand_totals["br_eth"] += br_eth
        grand_totals["br_fiat"] += br_fiat
        grand_totals["total_eth"] += total_eth
        grand_totals["total_fiat"] += total_fiat
        grand_totals["p_eth"] += p_eth
        grand_totals["p_fiat"] += p_fiat

    # Grand total row
    ws_summary.append([])
    total_row = [
        "TOTAL",
        grand_totals["w_eth"], grand_totals["w_fiat"],
        grand_totals["br_eth"], grand_totals["br_fiat"],
        grand_totals["total_eth"], grand_totals["total_fiat"],
    ]
    if has_principal:
        total_row.extend([grand_totals["p_eth"], grand_totals["p_fiat"]])
    ws_summary.append(total_row)
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = header_font

    # Format summary number columns (ETH = 9 decimals, fiat = 2 decimals)
    num_cols = 9 if has_principal else 7
    for row in ws_summary.iter_rows(min_row=2, min_col=2, max_col=num_cols):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0.000000000" if cell.column % 2 == 0 else "#,##0.00"

    # Auto-size summary columns
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_summary.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    # Validators sheet (when validator info is available)
    if validator_info:
        ws_val = wb.create_sheet(title="Validators", index=1)
        val_header = ["Index", "Account", "Type", "Status", "Balance [ETH]",
                      "Effective Balance [ETH]", "Withdrawn [ETH]", "Withdrawn Income [ETH]"]
        ws_val.append(val_header)
        for cell in ws_val[1]:
            cell.font = header_font

        total_balance = Decimal(0)
        total_effective = Decimal(0)
        total_withdrawn = Decimal(0)
        total_income = Decimal(0)

        for vi in sorted(validator_info, key=lambda x: x.get("index", 0)):
            cred = vi.get("credential_type", "0x01")
            type_label = "Compounding (0x02)" if cred == "0x02" else "Distributing (0x01)"
            bal = vi.get("balance_eth", Decimal(0))
            eff = vi.get("effective_balance_eth", Decimal(0))
            withdrawn = vi.get("withdrawn_eth", Decimal(0))
            income = vi.get("withdrawn_income_eth", Decimal(0))
            ws_val.append([
                vi.get("index", ""),
                vi.get("name", ""),
                type_label,
                vi.get("status", ""),
                bal, eff, withdrawn, income,
            ])
            total_balance += bal
            total_effective += eff
            total_withdrawn += withdrawn
            total_income += income

        # Totals row
        ws_val.append([])
        ws_val.append(["", "TOTAL", "", "", total_balance, total_effective, total_withdrawn, total_income])
        for cell in ws_val[ws_val.max_row]:
            cell.font = header_font

        # Format number columns
        for row in ws_val.iter_rows(min_row=2, min_col=5, max_col=8):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "0.000000000"

        for col in ws_val.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws_val.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    return wb


def write_excel(events: List[Dict], prices: List[Tuple[int, Decimal]], account_name: str, date_from: str, date_to: str,
                currency: str = "EUR", output_dir: str = ".", validator_info: Optional[List[Dict]] = None):
    """Write Excel workbook to disk."""
    wb = _build_excel_workbook(events, prices, currency, validator_info=validator_info)
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "", account_name) or "unnamed"
    filename = f"staking_report_{safe_name}_{date_from}_{date_to}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"\n  Saved: {filepath}")
    return filepath


def write_excel_bytes(events: List[Dict], prices: List[Tuple[int, Decimal]], currency: str = "EUR",
                      validator_info: Optional[List[Dict]] = None, tz_label: str = "UTC") -> bytes:
    """Generate Excel report in memory and return as bytes."""
    import io
    wb = _build_excel_workbook(events, prices, currency, validator_info=validator_info, tz_label=tz_label)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def gather_events(
    account: dict,
    etherscan: EtherscanClient,
    start_ts: int,
    end_ts: int,
    log_fn=None,
) -> List[Dict]:
    """Fetch all staking events for an account. Returns raw event list (no file I/O)."""
    log = log_fn or print
    name = account["name"]
    log(f"\nProcessing account: {name}")

    withdrawal_address = account.get("withdrawal_address", "")

    if not withdrawal_address:
        raise RuntimeError(
            f"Account '{name}' needs a withdrawal/fee recipient address. "
            "This is the address that receives your staking rewards."
        )

    # Step 1: Fetch withdrawals from Etherscan
    log("  Fetching withdrawals...")
    withdrawals = etherscan.get_beacon_withdrawals(withdrawal_address, start_ts, end_ts)

    # Filter withdrawals by validator index if specific validators were requested
    validator_indices = account.get("validator_indices")
    if validator_indices:
        validator_set = set(validator_indices)
        filtered = [w for w in withdrawals if w["validator"] in validator_set]
        log(f"  Filtered withdrawals: {len(filtered)}/{len(withdrawals)} match validator(s) {validator_indices}")
        withdrawals = filtered

    # Step 2: For 0x02 validators, classify withdrawals as income vs principal
    credential_type = account.get("credential_type", "0x01")
    validator_pubkeys = account.get("validator_pubkeys", {})  # {index: pubkey}

    if credential_type == "0x02" and validator_pubkeys:
        log("  0x02 compounding validator detected — classifying withdrawals...")

        # Fetch EIP-7002 withdrawal requests to identify user-initiated withdrawals
        withdrawal_requests = etherscan.get_withdrawal_requests(withdrawal_address, start_ts, end_ts)

        # Fetch EIP-7251 consolidation events to exclude consolidation transfers
        consolidations = etherscan.get_consolidation_events(withdrawal_address, start_ts, end_ts)
        consolidated_pubkeys = {c["source_pubkey"] for c in consolidations if not c["is_upgrade"]}

        income_count = 0
        principal_count = 0
        consolidated_count = 0

        for w in withdrawals:
            vid = w["validator"]
            pubkey = validator_pubkeys.get(vid, "").lower()

            # Check if this validator was consolidated (source of a merge)
            if pubkey and pubkey in consolidated_pubkeys:
                w["type"] = "withdrawal_principal"
                consolidated_count += 1
                continue

            # Check if there's a matching EIP-7002 withdrawal request
            if pubkey and pubkey in withdrawal_requests:
                w["type"] = "withdrawal_principal"
                principal_count += 1
            else:
                # No matching request = skimming (auto-sweep above cap) = income
                income_count += 1

        log(f"  Classification: {income_count} income (skimming), {principal_count} principal (user-requested), {consolidated_count} consolidation")

    # Step 3: Fetch execution rewards (local blocks + MEV) from Etherscan
    # Use fee_recipient address if provided (can differ from withdrawal address)
    fee_recipient = account.get("fee_recipient", "") or withdrawal_address
    log(f"  Fetching execution layer rewards (local blocks + MEV) for {fee_recipient[:10]}...{fee_recipient[-4:]}...")
    block_rewards = etherscan.get_execution_rewards(fee_recipient, start_ts, end_ts)

    # Tag block rewards with validator index when a single validator is queried
    if validator_indices and len(validator_indices) == 1:
        for br in block_rewards:
            br["validator"] = validator_indices[0]

    return withdrawals + block_rewards


def process_account(
    account: dict,
    etherscan: EtherscanClient,
    prices: List[Tuple[int, Decimal]],
    start_ts: int,
    end_ts: int,
    date_from: str,
    date_to: str,
    currency: str = "EUR",
    output_format: str = "xlsx",
    output_dir: str = ".",
):
    """Fetch events and write report files to disk (used by CLI)."""
    all_events = gather_events(account, etherscan, start_ts, end_ts)

    if not all_events:
        print(f"  No events found for account '{account['name']}' in date range.")
        return None

    files = []
    if output_format in ("xlsx", "both"):
        print("  Writing Excel report...")
        files.append(write_excel(all_events, prices, account["name"], date_from, date_to, currency, output_dir))
    if output_format in ("csv", "both"):
        print("  Writing CSV report...")
        files.append(write_csv(all_events, prices, account["name"], date_from, date_to, currency, output_dir))
    return files


def main():
    parser = argparse.ArgumentParser(description="Ethereum staking tax report generator")
    parser.add_argument("--from", dest="date_from", required=True, help="Start date (YYYY-MM-DD)")
    parser.add_argument("--to", dest="date_to", required=True, help="End date (YYYY-MM-DD)")
    parser.add_argument("--account", help="Process only this account name")
    parser.add_argument("--currency", default="EUR", choices=[c.lower() for c in SUPPORTED_CURRENCIES], help="Fiat currency (default: EUR)")
    parser.add_argument("--format", dest="output_format", default="xlsx", choices=["xlsx", "csv", "both"], help="Output format (default: xlsx)")
    parser.add_argument("--config", default="config.yaml", help="Path to config file")
    args = parser.parse_args()

    config = load_config(args.config)
    currency = args.currency.upper()

    # Parse dates
    start_dt = datetime.strptime(args.date_from, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end_dt = datetime.strptime(args.date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    start_ts = int(start_dt.timestamp())
    end_ts = int(end_dt.timestamp())

    print(f"Date range: {args.date_from} to {args.date_to}")
    print(f"Currency: {currency}")

    # Init clients
    etherscan = EtherscanClient(config["etherscan_api_key"])
    price_client = CryptoCompareClient(currency=currency)
    print(f"Fetching ETH/{currency} prices...")
    prices = price_client.get_prices(start_ts, end_ts)

    # Process accounts
    accounts = config.get("accounts", [])
    if args.account:
        accounts = [a for a in accounts if a["name"] == args.account]
        if not accounts:
            print(f"Account '{args.account}' not found in config.")
            sys.exit(1)

    # Resolve validator index/pubkey to withdrawal address where needed
    beacon = None
    for account in accounts:
        if "validator_indices" in account:
            # List of validator indices in config
            if beacon is None:
                beacon = BeaconClient()
            indices = []
            pubkeys = {}
            address = None
            cred_type = "0x01"
            for vid in account["validator_indices"]:
                info = beacon.get_validator_info(str(vid))
                indices.append(info["validator_index"])
                pubkeys[info["validator_index"]] = info.get("pubkey", "")
                cred_type = info.get("credential_type", "0x01")
                if address is None:
                    address = info["withdrawal_address"]
            account["withdrawal_address"] = address
            account["validator_indices"] = indices
            account["credential_type"] = cred_type
            account["validator_pubkeys"] = pubkeys
        elif "validator_index" in account:
            if beacon is None:
                beacon = BeaconClient()
            info = beacon.get_validator_info(str(account["validator_index"]))
            account["withdrawal_address"] = info["withdrawal_address"]
            account["validator_indices"] = [info["validator_index"]]
            account["credential_type"] = info.get("credential_type", "0x01")
            account["validator_pubkeys"] = {info["validator_index"]: info.get("pubkey", "")}
        elif "validator_pubkey" in account:
            if beacon is None:
                beacon = BeaconClient()
            info = beacon.get_validator_info(account["validator_pubkey"])
            account["withdrawal_address"] = info["withdrawal_address"]
            account["validator_indices"] = [info["validator_index"]]
            account["credential_type"] = info.get("credential_type", "0x01")
            account["validator_pubkeys"] = {info["validator_index"]: info.get("pubkey", "")}

    for account in accounts:
        process_account(
            account=account,
            etherscan=etherscan,
            prices=prices,
            start_ts=start_ts,
            end_ts=end_ts,
            date_from=args.date_from,
            date_to=args.date_to,
            currency=currency,
            output_format=args.output_format,
        )

    print("\nDone!")


if __name__ == "__main__":
    main()
